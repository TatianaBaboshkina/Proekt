import openpyxl
from tkinter import *

wb_obj = openpyxl.load_workbook("oko_2.xlsx")
sheet_obj = wb_obj.active

#m - максимальное количество строк
#c - максимальное количество столбов
#ssred - сумма показателей по строке
#sred - среднее значение
#b - счетчик количества слагаемых
m = sheet_obj.max_row
c = sheet_obj.max_column
pokazatel = 0
m = m+1 
i = 2
j= 2
c = sheet_obj.max_column
ssred = 0
sred=0
b=0
o = 0
window = Tk()
window.title("Удовлетворённость клиентов ДЦ ГАЗ")
while j < c and i<m:
    znach = sheet_obj.cell(row = i, column = j)
    pokazatel = sheet_obj.cell(row = i, column = 1)
    if znach.value!=None:
        ssred = ssred + znach.value
        j=j+1
        b = b + 1
    else:           
        j = 2
        i = i + 1
        sred = ssred / b
        ssred = 0
        b = 0
        print(str(pokazatel.value) + " " + str(sred))
        lbl = Label(window, text = str(pokazatel.value) + " - " + str(sred))   
        lbl.grid(column=0, row=o)
        o = o + 1
window.mainloop()

    
